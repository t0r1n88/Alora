"""
Скрипт для подсчета свода по районам и школам среди тех кто начал обучение и прошел хотя бы один тест
"""

import pandas as pd
import time
import os
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def convert_to_none(cell):
    """
    Для замены дефисов на нан
    :param cell:
    :return:
    """
    if isinstance(cell,str):
        if cell == '-':
            return None
        else:
            return cell
    else:
        return cell


def calc_first_part(row):
    """
    Функция для подсчета сдавших первый модуль
    :param row:
    :return:
    """
    first_task,second_task = row # Задание и тест 3 и 5
    if first_task >= 3.00 and second_task >= 5.00:
        return 1
    else:
        return 0

def calc_second_part(row):
    """
    Функция для подсчета сдавших второй модуль
    :param row:
    :return:
    """
    if all(x > 3.00 for x in row):
        return 1
    else:
        return 0


def calc_third_part(row):
    """
    Функция для подсчета сдавших третий модуль
    :param row:
    :return:
    """
    if all(x > 3.00 for x in row[:4]) and row[4] >= 3.00 and row[5] >= 5.00:
        return 1
    else:
        return 0


def create_lists(df:pd.DataFrame,result_folder:str,value_filter:int,name_column:str,name_folder:str):
    """
    Функция для сохранения списков сдавших и не сдавших модули
    :param df: датафрейм с данными
    :param result_folder: конечная папка
    :param value_filter: значение 0 или 1 по которому определяется, сдал школьник модуль или нет
    :param name_column: название колонки где указано сдал или не сдал модуль
    :param name_folder: название папки которую нужно создать
    """
    df = df[df['Муниципалитет'].notna()]
    # Сохраняем списки по муниципалитетам
    lst_value_column = df['Муниципалитет'].unique()

    for idx, value in enumerate(lst_value_column):
        wb = openpyxl.Workbook()  # создаем файл
        temp_df = df[df['Муниципалитет'] == value]  # отфильтровываем по значению
        temp_df = temp_df[temp_df[name_column]==value_filter]
        temp_df.rename(columns={'сдавших первый модуль':'Первый_модуль','сдавших второй модуль':'Второй_модуль','сдавших третий модуль':'Третий_модуль'},inplace=True)
        temp_df[['Первый_модуль','Второй_модуль','Третий_модуль']] = temp_df[['Первый_модуль','Второй_модуль','Третий_модуль']].replace({0:'не сдан',1:'сдан'})


        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb['Sheet'].append(row)

        # Устанавливаем автоширину для каждой колонки
        for column in wb['Sheet'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['Sheet'].column_dimensions[column_name].width = adjusted_width

        finish_path = f'{result_folder}/{name_folder}'
        if not os.path.exists(finish_path):
            os.makedirs(finish_path)

        wb.save(f'{finish_path}/{value} учеников-{len(temp_df)}.xlsx')
        wb.close()







def generate_svod_first_prof(list_student:str,estimation_file:str,lst_moodle:str,result_folder:str):
    """
    Функция для создания свод по прошедшим обучение.
    :param list_student: итоговый список школьников
    :param estimation_file: файл с оценками
    :param lst_moodle: файл логинами мудл
    :param result_folder: конечная папка
    :return:
    """
    try:
        # получаем время
        t = time.localtime()
        current_time = time.strftime('%H_%M на %d.%m', t)
        df = pd.read_excel(list_student,dtype=str) # файл с данными школьников
        est_df = pd.read_excel(estimation_file)
        est_df.drop(columns=['Индивидуальный номер','Учреждение (организация)','Отдел','Адрес электронной почты','Последние загруженные из этого курса'],inplace=True)

        est_df = est_df.applymap(convert_to_none)

        # Создаем колонки с ФИО
        df['ФИО'] = df['Фамилия'] + ' ' + df['Имя'] + ' '+ df['Отчество']
        est_df['ФИО'] = est_df['Фамилия'] + ' ' + est_df['Имя']

        df = pd.merge(df,est_df,how='outer',left_on='ФИО',right_on='ФИО')

        not_start_df = df[df['Тест:Тест 1.1. Речевая и логическая культура ведения делового разговора (Значение)'].isna()]
        not_start_df = not_start_df[['Школа','Класс','ФИО_представителя','Номер_телефона','Электронная_почта','ФИО','Муниципалитет']]

        # Соединяем с файлом мудла
        moodle_df = pd.read_excel(lst_moodle,dtype=str)
        moodle_df['ФИО'] = moodle_df['lastname'] + ' ' + moodle_df['firstname']

        not_start_df = pd.merge(not_start_df,moodle_df,how='inner',left_on='ФИО',right_on='ФИО')
        not_start_df.drop(columns=['firstname','lastname','email','cohort1'],inplace=True)
        not_start_df.rename(columns={'ФИО':'ФИО школьника','username':'Логин для edu-copp03.ru','password':'Пароль для edu-copp03.ru'},inplace=True)
        not_start_df['Телеграм курса'] = 'https://t.me/+Zgw_U0s4hCUzMTNi'
        not_start_df=not_start_df.reindex(columns=['ФИО_представителя','Номер_телефона','Электронная_почта','ФИО школьника','Логин для edu-copp03.ru','Пароль для edu-copp03.ru','Телеграм курса',
                                      'Муниципалитет','Школа','Класс'])

        # Сохраняем списки по муниципалитетам
        lst_value_column = not_start_df['Муниципалитет'].unique()
        finish_path = f'{result_folder}/НЕ ПРИСТУПИВШИЕ К ОБУЧЕНИЮ'
        if not os.path.exists(finish_path):
            os.makedirs(finish_path)

        for idx, value in enumerate(lst_value_column):
            wb = openpyxl.Workbook()  # создаем файл
            temp_df = not_start_df[not_start_df['Муниципалитет'] == value]  # отфильтровываем по значению
            # temp_df = temp_df[['Школа','Класс','ФИО школьника']]
            for row in dataframe_to_rows(temp_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Устанавливаем автоширину для каждой колонки
            for column in wb['Sheet'].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb['Sheet'].column_dimensions[column_name].width = adjusted_width

            wb.save(f'{finish_path}/{value} учеников-{len(temp_df)}.xlsx')
            wb.close()

        # Считаем прошедших модули
        df['сдавших первый модуль'] = df[['Задание:Задание самостоятельной работы 1.2.1 (Значение)','Тест:Тест 1.1. Речевая и логическая культура ведения делового разговора (Значение)']].apply(calc_first_part,axis=1)
        df['сдавших второй модуль'] = df[['Задание:Задание 2.1 (Значение)','Задание:Задание 2.2.1 (Значение)','Задание:Задание 2.2.2 (Значение)','Задание:Задание 2.2.3 (Значение)']].apply(calc_second_part,axis=1)
        df['сдавших третий модуль'] = df[['Задание:Задание 3.2. (Значение)','Задание:Задание 3.6.1 (Значение)','Задание:Задание 3.6.2 (Значение)','Задание:Задание 3.6.3 (Значение)',
                                          'Тест:Тест 3.3. Поисковые системы в сети Интернет: принципы работы и функциональные возможности (Значение)',
                                          'Тест:Тест 3.7. Основы работы с базами данных (Значение)']].apply(calc_third_part,axis=1)


        """
        Создаем списки сдавших и не сдавших каждый модуль
        """
        lst_for_lists = ['ФИО_представителя','Номер_телефона','Электронная_почта','ФИО',
                                      'Муниципалитет','Школа','Класс','сдавших первый модуль','сдавших второй модуль','сдавших третий модуль']

        # 1 модуль
        create_lists(df[lst_for_lists].copy(),result_folder,1,'сдавших первый модуль','Сдавшие ПЕРВЫЙ модуль')
        create_lists(df[lst_for_lists].copy(),result_folder,0,'сдавших первый модуль','НЕ сдавшие ПЕРВЫЙ модуль')

        # 2 модуль
        create_lists(df[lst_for_lists].copy(),result_folder,1,'сдавших второй модуль','Сдавшие ВТОРОЙ модуль')
        create_lists(df[lst_for_lists].copy(),result_folder,0,'сдавших второй модуль','НЕ сдавшие ВТОРОЙ модуль')

        # 3 модуль
        create_lists(df[lst_for_lists].copy(),result_folder,1,'сдавших третий модуль','Сдавшие ТРЕТИЙ модуль')
        create_lists(df[lst_for_lists].copy(),result_folder,0,'сдавших третий модуль','НЕ сдавшие ТРЕТИЙ модуль')




        group_df = df.groupby(['Муниципалитет']).agg({'Тест:Тест 1.1. Речевая и логическая культура ведения делового разговора (Значение)':'count','сдавших первый модуль':'sum',
                                                      'сдавших второй модуль':'sum','сдавших третий модуль':'sum'}).fillna(0)
        group_df.rename(columns={'Тест:Тест 1.1. Речевая и логическая культура ведения делового разговора (Значение)':'приступивших к обучению'},inplace=True)
        # Добавляем колонку с количеством зарегистрировавшихся
        group_df.insert(0,'Записано на курс',[6,7,12,10,7,23,32,22,20,8,30,23,34,10,10,31,8,4,38])

        # Приступившие
        group_df['% приступивших к обучению'] = round((group_df['приступивших к обучению'] / group_df['Записано на курс']) * 100,1)
        group_df['НЕ приступивших к обучению'] = group_df['Записано на курс'] - group_df['приступивших к обучению']
        # Первый модуль
        group_df['% сдавших первый модуль'] = round((group_df['сдавших первый модуль'] / group_df['Записано на курс']) * 100,1)
        # Второй модуль
        group_df['% сдавших второй модуль'] = round((group_df['сдавших второй модуль'] / group_df['Записано на курс']) * 100, 1)
        # Третий модуль
        group_df['% сдавших третий модуль'] = round((group_df['сдавших третий модуль'] / group_df['Записано на курс']) * 100, 1)



        group_df = group_df.reindex(columns=['Записано на курс','приступивших к обучению','% приступивших к обучению','НЕ приступивших к обучению','сдавших первый модуль','% сдавших первый модуль',
                                  'сдавших второй модуль','% сдавших второй модуль', 'сдавших третий модуль','% сдавших третий модуль'])


        sum_row = group_df.sum(axis=0, numeric_only=True)
        sum_row = sum_row.rename('Итого').to_frame().transpose()
        group_df = pd.concat([group_df, sum_row])
        group_df.loc['Итого', '% приступивших к обучению'] = round(
            (group_df.loc['Итого', 'приступивших к обучению'] / group_df.loc['Итого', 'Записано на курс']) * 100, 1)

        group_df.loc['Итого', '% сдавших первый модуль'] = round(
            (group_df.loc['Итого', 'сдавших первый модуль'] / group_df.loc['Итого', 'Записано на курс']) * 100, 1)

        group_df.loc['Итого', '% сдавших второй модуль'] = round(
            (group_df.loc['Итого', 'сдавших второй модуль'] / group_df.loc['Итого', 'Записано на курс']) * 100, 1)

        group_df.loc['Итого', '% сдавших третий модуль'] = round(
            (group_df.loc['Итого', 'сдавших третий модуль'] / group_df.loc['Итого', 'Записано на курс']) * 100, 1)

        with pd.ExcelWriter(f'{result_folder}/Сводка Первая профессия в {current_time}.xlsx') as writer:
            group_df.to_excel(writer, sheet_name='Свод по муниципалитетам', index=True)
    except PermissionError as e:
        messagebox.showerror('Алора',
                             f'Закройте файлы созданные программой')
    except FileNotFoundError as e:
        messagebox.showerror('Алора',
                             f'Не удалось создать файл с названием {e}\n'
                             f'Уменьшите символов в соответствующей строке или выберите более короткий путь к итоговой папке')
    else:
        messagebox.showinfo('Алора', 'Создание документов успешно завершено !')


if __name__ == '__main__':
    main_list_student = 'data/ИТОГОВЫЙ список зарегистрировавшихся на курс.xlsx'
    main_estimation_file = 'data/Цифровой куратор Оценки.xlsx'
    main_lst_moodle = 'data/Файл для MOODLE 27_10.xlsx'
    main_result_folder = 'data/Результат'

    generate_svod_first_prof(main_list_student,main_estimation_file,main_lst_moodle,main_result_folder)

    print('Lindy Booth')