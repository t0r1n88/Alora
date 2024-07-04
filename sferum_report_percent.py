"""
Скрипт для обработки данных мониторинга уровня использования Сферум по районам
"""
import pandas as pd
import openpyxl
from tkinter import messagebox
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import time
import re

class NoNameColumn(BaseException):
    """
    Класс для обозначения ситуации когда не найдена колонка с нужным именем
    """
    pass

"""
Алгоритм
1) Отфильтровать данные по республике
2) Создать документ openpyxl
3) Отфильтровать данные по району и создать лист с соответствующим названием
4) Поместить данные на этот лист и закрасить красным строки где в определенных колонках значение меньше целевого показателя

Примечания
Размер целевого показателя должен вводиться через поле ввода  на данный момент 32,63%
Названия листа с которого будут забираться данные тоже должны вводится
"""
def process_split_highlighting_threshold(path:str,name_sheet:str,region:str,path_to_end_folder:str,threshold)-> None:
    try:
        df = pd.read_excel(path,dtype={'ИНН':str})
        lst_need_columns = ['id','Субъект РФ','Наименование АТО','Краткое наименование организации',
                            'Полное наименование организации','Адрес организации','ИНН','Тип поселения',
                            'Статус организации\n(юр.лицо/филиал)','Статус\nфункционирования',
                            'Государственная/негосударственная организация (ГОУ/НОУ)',
                            'Численность обучающихся\nвсего, чел. (Раздел 1.3. стр.01 гр.3)',
                            'Численность педработников - всего, чел. (Раздел 3.1 стр.06 гр.3)',
                            'Обучающиеся Сферум (чел)','Обучающиеся МШ (чел)','Обучающиеся Cферум или МШ (чел)',
                            'Обучающиеся, использовавшие за последние 12 месяцев Сферум',
                            ' Педагоги, использовавшие за последние 12 месяцев Сферум','Обучающиеся, использовавшие за последние 12 месяцев МШ',
                            ' Педагоги, использовавшие за последние 12 месяцев МШ','Педагоги, использовавшие за последние 12 месяцев Сферум или МШ (чел)',
                            'Обучающиеся, использовавшие за последние 12 месяцев Сферум % от общего числа',
                            'Педагоги, использовавшие за последние 12 месяцев Сферум % от общего числа',
                            'Обучающиеся, использовавшие за последние 12 месяцев МШ % от общего числа',
                            'Педагоги, использовавшие за последние 12 месяцев МШ % от общего числа']

        diff_set = set(lst_need_columns).difference(set(df.columns))
        if len(diff_set) > 0 :
            raise NoNameColumn

        # Отбираем нужные колонки
        df = df[lst_need_columns]
        df = df[df['Субъект РФ'] == region] # Фильтруем регион
        lst_target_column = ['Обучающиеся, использовавшие за последние 12 месяцев Сферум % от общего числа',
                             'Педагоги, использовавшие за последние 12 месяцев Сферум % от общего числа',
                             'Обучающиеся, использовавшие за последние 12 месяцев МШ % от общего числа',
                             'Педагоги, использовавшие за последние 12 месяцев МШ % от общего числа',]
        # Округляем значения в последних 4 колонках
        df[lst_target_column] = df[lst_target_column].apply(lambda x:round(x,4)*100)

        lst_unqiue_ruo = df['Наименование АТО'].unique() # получаем список РУО

        used_name_sheet = set()  # множество для хранения значений которые уже были использованы
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Создаем документ и листы
        wb = openpyxl.Workbook()
        for idx,name_ruo in enumerate(lst_unqiue_ruo):
            temp_df = df[df['Наименование АТО'] == name_ruo]

            short_value = name_ruo[:20]  # получаем обрезанное значение
            short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

            if short_value.lower() in used_name_sheet:
                short_value = f'{short_value}_{idx}'  # добавляем окончание
            wb.create_sheet(short_value, index=idx)  # создаем лист
            used_name_sheet.add(short_value.lower())
            for row in dataframe_to_rows(temp_df, index=False, header=True):

                wb[short_value].append(row)
                # Форматируем документ
                wb[short_value]['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['B1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['C1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['D1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value].column_dimensions['D'].width = 30
                wb[short_value]['E1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['F1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['G1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['H1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['I1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['J1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['K1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['L1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['M1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['O1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['P1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['Q1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['R1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['S1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['T1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['U1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['V1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['W1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['X1'].alignment = openpyxl.styles.Alignment(wrapText=True)
                wb[short_value]['Y1'].alignment = openpyxl.styles.Alignment(wrapText=True)

                lst_index_col = []
                for name_column in lst_target_column:
                    lst_index_col.append(list(temp_df.columns).index(name_column))

                for ind_col in lst_index_col:
                    for row in wb[short_value].iter_rows(min_row=2,max_row=wb[short_value].max_row,
                                                         min_col=ind_col+1,max_col=ind_col+1):
                        for cell in row:
                            if float(cell.value) < threshold:
                                cell.fill = PatternFill(fill_type='solid', fgColor='ff0000')



            wb.save(f'{path_to_end_folder}/Районы МШ и Сферум {current_time}.xlsx')
            wb.close()
    except ValueError as e:
        messagebox.showerror('',f'Обнаружена ошибка {e.args}')
    else:
        messagebox.showinfo('','Обработка завершена успешно')



if __name__ == '__main__':
    main_df = 'data/Расчет_2.1_2.2_2.3_июнь_2024 г.xlsx'
    main_name_sheet = 'Процент пользователей'
    main_region = 'Республика Бурятия'
    main_region = 'Приморский край'
    main_end_folder = 'data/Результат'
    main_threshold = 32.63
    process_split_highlighting_threshold(main_df, main_name_sheet, main_region,main_end_folder,main_threshold)

    print('Lindy Booth')


