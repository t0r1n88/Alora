"""
Функции  для нахождения разницы двух таблиц
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from tkinter import messagebox
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)


# Классы для исключений
class ShapeDiffierence(Exception):
    """
    Класс для обозначения несовпадения размеров таблицы
    """
    pass


class ColumnsDifference(Exception):
    """
    Класс для обозначения того что названия колонок не совпадают
    """
    pass


def write_df_to_excel(dct_df:dict,write_index:bool)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        if len(df) == 0:
            continue
        for row in dataframe_to_rows(df,index=write_index,header=True):
            wb[name_sheet].append(row)
        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
    # удаляем лишний лист
    if len(wb.sheetnames) >= 2 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb






def abs_diff(first_value, second_value):
    """
    Функция для подсчета абсолютной разницы между 2 значениями
    """
    try:
        return abs(float(first_value) - float(second_value))
    except:
        return None


def percent_diff(first_value, second_value):
    """
    функция для подсчета относительной разницы значений
    """
    try:
        # округляем до трех
        value = round(float(second_value) / float(first_value), 4) * 100
        return value
    except:
        return None


def change_perc_diff(first_value, second_value):
    """
    функция для подсчета процентного ихменения значений
    """
    try:
        value = (float(second_value) - float(first_value)) / float(first_value)
        return round(value, 4) * 100
    except:
        return None


def find_diffrence(first_sheet, second_sheet, first_df, second_df,path_to_end_folder_diffrence):
    """
    Функция для вычисления разницы между 2 таблицами
    :param first_sheet: Имя первого листа
    :param second_sheet: Имя второго листа
    :param first_df: Путь к первой таблице
    :param second_df: Путь ко второй таблице
    :param path_to_end_folder_diffrence : Путь к папке куда будут сохранятся файлы
    :return: разница между двумия таблица файл Excel в котором 3 листа:
    По колонкам - в котором указаны те ячейки в которых найдена разница
    По строкам - тоже самое только отображение по строкам
    Значение разницы - таблица с указанием абсолютной и процентной разницы между измененными значениями
    """

    # загружаем датафреймы
    try:
        try:
            df1 = pd.read_excel(first_df, sheet_name=first_sheet, dtype=str)
            df2 = pd.read_excel(second_df, sheet_name=second_sheet, dtype=str)
        except ValueError:
            messagebox.showerror('Алора',
                                 f'В файлах нет листа с таким названием!\n'
                                 f'Проверьте написание названия листа')
        except:
            messagebox.showerror('Алора',
                                 f'Не удалось обработать файлы . Возможно какой то из файлов используемых для сравнения поврежден')

        # проверяем на соответсвие размеров
        if df1.shape != df2.shape:
            raise ShapeDiffierence

        # Проверям на соответсвие колонок
        if list(df1.columns) != list(df2.columns):
            diff_columns = set(df1.columns).difference(set(df2.columns))  # получаем отличающиеся элементы
            raise ColumnsDifference

        df_cols = df1.compare(df2,
                              result_names=('Первая таблица', 'Вторая таблица'))  # датафрейм с разницей по колонкам
        df_cols.index = list(
            map(lambda x: x + 2, df_cols.index))  # добавляем к индексу +2 чтобы соответствовать нумерации в экселе
        df_cols.index.name = '№ строки'  # переименовываем индекс

        df_rows = df1.compare(df2, align_axis=0,
                              result_names=('Первая таблица', 'Вторая таблица'))  # датафрейм с разницей по строкам
        lst_mul_ind = list(map(lambda x: (x[0] + 2, x[1]),
                               df_rows.index))  # добавляем к индексу +2 чтобы соответствовать нумерации в экселе
        index = pd.MultiIndex.from_tuples(lst_mul_ind, names=['№ строки', 'Таблица'])  # создаем мультиндекс
        df_rows.index = index

        # Создаем датафрейм с подсчетом разниц
        df_diff_cols = df_cols.copy()

        # получаем список колонок первого уровня
        temp_first_level_column = list(map(lambda x: x[0], df_diff_cols.columns))
        first_level_column = []
        [first_level_column.append(value) for value in temp_first_level_column if value not in first_level_column]

        # Добавляем колонки с абсолютной и относительной разницей
        count_columns = 2
        for name_column in first_level_column:
            # высчитываем абсолютную разницу
            df_diff_cols.insert(count_columns, (name_column, 'Разница между первым и вторым значением'),
                                df_diff_cols.apply(lambda x: abs_diff(x[name_column]['Первая таблица'],
                                                                      x[name_column]['Вторая таблица']), axis=1))

            # высчитываем отношение второго значения от первого
            df_diff_cols.insert(count_columns + 1, (name_column, '% второго от первого значения'),
                                df_diff_cols.apply(lambda x: percent_diff(x[name_column]['Первая таблица'],
                                                                          x[name_column]['Вторая таблица']), axis=1))

            # высчитываем процентное изменение
            df_diff_cols.insert(count_columns + 2, (name_column, 'Изменение в процентах'),
                                df_diff_cols.apply(lambda x: change_perc_diff(x[name_column]['Первая таблица'],
                                                                              x[name_column]['Вторая таблица']),
                                                   axis=1))

            count_columns += 5

        # записываем
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # записываем в файл Excel с сохранением ширины
        dct_df = {'По колонкам':df_cols,'По строкам':df_rows,'Значение разницы':df_diff_cols}
        write_index = True # нужно ли записывать индекс
        wb = write_df_to_excel(dct_df,write_index)
        wb.save(f'{path_to_end_folder_diffrence}/Разница между 2 таблицами {current_time}.xlsx')
    except UnboundLocalError:
        pass
    except ShapeDiffierence:
        messagebox.showerror('Алора',
                             f'Не совпадают размеры таблиц, В первой таблице {df1.shape[0]}-стр. и {df1.shape[1]}-кол.\n'
                             f'Во второй таблице {df2.shape[0]}-стр. и {df2.shape[1]}-кол.')

    except ColumnsDifference:
        messagebox.showerror('Алора',
                             f'Названия колонок в сравниваемых таблицах отличаются\n'
                             f'Колонок:{diff_columns}  нет во второй таблице !!!\n'
                             f'Сделайте названия колонок одинаковыми.')

    except ValueError:
        messagebox.showerror('Алора',
                             f'В файлах нет листа с таким названием!\n'
                             f'Проверьте написание названия листа')
        logging.exception('AN ERROR HAS OCCURRED')
    except FileNotFoundError:
        messagebox.showerror('Алора',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')

    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Алора',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log Возможно обрабатываемые файлы повреждены.')
    else:
        messagebox.showinfo('Алора', 'Таблицы успешно обработаны')

if __name__ == '__main__':
    first_sheet_main = 'Основное'
    second_sheet_main = 'Основное'
    data_first_diffrence_main = 'data\Разница между 2 таблицами\Отчет 2021.xlsx'
    data_second_diffrence_main = 'data\Разница между 2 таблицами\Отчет 2022.xlsx'
    data_second_diffrence_main = 'data\Разница между 2 таблицами\Тест.xlsx'
    path_to_end_folder_diffrence_main = 'data'

    find_diffrence(first_sheet_main, second_sheet_main, data_first_diffrence_main, data_second_diffrence_main, path_to_end_folder_diffrence_main)

    print('Lindy Booth')