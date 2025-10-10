"""
Скрипты для обработки таблицы с данными по первой профессии
"""
import pandas as pd
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import os
import random
import time
from tkinter import messagebox
from docx.opc.exceptions import PackageNotFoundError

def capitalize_fio(value:str)->str:
    """
    Функция для применения capitalize к значениям состоящим из несколько слов разделенных пробелами
    value: значение ячейки
    """
    value = str(value)
    if value == 'Не заполнено':
        return value
    temp_lst = value.split(' ') # создаем список по пробелу
    temp_lst = list(map(str.capitalize,temp_lst))  # обрабатываем
    return ' '.join(temp_lst) #соединяем в строку


def prepare_fio_text_columns(df:pd.DataFrame,prepared_columns_lst:list)->pd.DataFrame:
    """
    Функция для очистки текстовых колонок c данными ФИО
    df: датафрейм для обработки
    lst_columns: список колонок которые нужно обработать
    """

    df[prepared_columns_lst] = df[prepared_columns_lst].fillna('Ошибка: Не заполнено')
    df[prepared_columns_lst] = df[prepared_columns_lst].astype(str)
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x: x.strip() if isinstance(x, str) else x)  # применяем strip, чтобы все данные корректно вставлялись
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x:' '.join(x.split())) # убираем лишние пробелы между словами
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(capitalize_fio)  # делаем заглавными первые буквы слов а остальыне строчными

    return df


def check_snils(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if snils is pd.isna(snils):
        return 'Ошибка Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр - {snils} -{len(result)} цифр'



def check_series_passport(series:str)->str:
    """
    Функция для проверки серии паспорта, должно быть 4 цифры
    """
    if series is pd.isna(series):
        return 'Ошибка: Не заполнено'
    series = str(series)
    result = re.findall(r'\d', series) # ищем цифры
    if len(result) == 4:
        return ''.join(result)
    else:
        return f'Ошибка: в серии паспорта должно быть 4 цифры - {series}'

def check_number_passport(number:str)->str:
    """
    Функция для проверки номера паспорта, должно быть 6 цифр
    """
    if number is pd.isna(number):
        return 'Ошибка: Не заполнено'
    number = str(number)
    result = re.findall(r'\d', number) # ищем цифры
    if len(result) == 6:
        return ''.join(result)
    else:
        return f'Ошибка: в номере паспорта должно быть 6 цифр - {number}'



def check_mixing(value:str):
    """
    Функция для проверки слова на смешение алфавитов
    """
    # ищем буквы русского и английского алфавита
    russian_letters = re.findall(r'[а-яА-ЯёЁ]',value)
    english_letters = re.findall(r'[a-zA-Z]',value)
    # если найдены и те и те
    if russian_letters and english_letters:
        # если русских букв больше то указываем что в русском слове встречаются английские буквы
        if len(russian_letters) > len(english_letters):
            return (f'Ошибка: в слове {value} найдены английские буквы: {",".join(english_letters)}')
        elif len(russian_letters) < len(english_letters):
            # если английских букв больше то указываем что в английском слове встречаются русские буквы
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)}')
        else:
            # если букв поровну то просто выводим их список
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)} и английские буквы: {";".join(english_letters)}')
    else:
        # если слово состоит из букв одного алфавита
        return False


def find_mixing_alphabets(cell):
    """
    Функция для нахождения случаев смешения когда английские буквы используются в русском слове и наоборот
    """
    if isinstance(cell,str):
        lst_word = re.split(r'\W',cell) # делим по не буквенным символам
        lst_result = list(map(check_mixing,lst_word)) # ищем смешения
        lst_result = [value for value in lst_result if value] # отбираем найденые смешения если они есть
        if lst_result:
            return f'Ошибка: в тексте {cell} найдено смешение русского и английского: {"; ".join(lst_result)}'
        else:
            return cell
    else:
        return cell

def write_df_to_excel_cheking_egisso(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            # для слишком длинных результатов
            if adjusted_width > 60:
                adjusted_width = 60
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

    return wb


def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb


def write_df_error_egisso_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма c данными ЕГИССО с заливкой ошибок цветом
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    for name_sheet, df in dct_df.items():
        short_name_sheet = name_sheet[:20]  # получаем обрезанное значение
        short_name_sheet = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_name_sheet)
        if short_name_sheet.lower() in used_name_sheet:
            short_name_sheet = f'{short_name_sheet}_{count_index}'  # добавляем окончание

        wb.create_sheet(title=short_name_sheet, index=count_index)  # создаем лист
        used_name_sheet.add(short_name_sheet.lower()) # добавляем в список использованных названий
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[short_name_sheet].append(row)
            else:
                wb[short_name_sheet].append(row)
        if none_check:
            wb[short_name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[short_name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 40:
                adjusted_width = 40
            wb[short_name_sheet].column_dimensions[column_name].width = adjusted_width



        count_index += 1

        start_column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[short_name_sheet].iter_rows(min_row=1, max_row=wb[short_name_sheet].max_row,
                                            min_col=start_column_number, max_col=df.shape[1]):  # Перебираем строки
            for i in range(1,df.shape[1]):
                if 'Ошибка' in str(row[i].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                    for cell in row:  # применяем стиль если условие сработало
                        cell.font = font
                        cell.fill = fill

    return wb


def create_cohort(value:str):
    """
    Функция для создания названия группы
    :param value: значение
    :return:
    """
    if value == 'г. Северобайкальск':
        return f'ПП25_Северобайкальск'
    elif value == 'Баунтовский (Эвенкийский) район':
        return f'ПП25_Баунтовский'
    else:
        name= value.split(' ')[0]
        return f'ПП25_{name}'








def processing_data_first_prof(path_to_data:str,result_folder:str):
    """
    :param begin_df: Путь к файлу с данными
    :param result_folder: Конечная папка
    :return:
    """
    try:

        # получаем время
        t = time.localtime()
        current_time = time.strftime('%H_%M на %d.%m', t)

        begin_df = pd.read_excel(path_to_data,dtype=str)
        number_mun = list(begin_df.columns).index('Выберите свой муниципалитет (район)') # Индекс колонки с муниципалитетом
        column_shool = list(begin_df.columns).index('Введите свою школу') # Индекс колонки с муниципалитетом

        lst_school = []

        for row in begin_df.iloc[:,number_mun+1:column_shool+1].itertuples():
            tmp_lst = list(row[1:])
            out_value = [value for value in tmp_lst if pd.notna(value)][0]
            lst_school.append(out_value)


        begin_df['Школа'] = lst_school
        # Начинаем собирать итоговый датафрейм
        df = begin_df[['Выберите свой муниципалитет (район)','Школа','Класс','Фамилия обучающегося','Имя обучающегося','Отчество обучающегося(при наличии)',
                       'Дата рождения обучающегося','Пол обучающегося','СНИЛС обучающегося','Фото СНИЛС обучающегося',
                       'Контактный телефон обучающегося','Гражданство обучающегося','Введите гражданство','Серия паспорта обучающегося',
                       'Номер паспорта обучающегося','Кем выдан паспорт обучающегося','Дата выдачи паспорта обучающегося',
                       'ФИО законного представителя','Серия паспорта законного представителя','Номер паспорта законного представителя',
                       'Кем выдан паспорт законного представителя','Дата выдачи паспорта законного представителя','Номер и серия свидетельства о рождении обучающегося',
                       'Дата выдачи свидетельства о рождении обучающегося','Сведения_об_ОВЗ','Номер телефона законного представителя','Электронная почта законного представителя']]


        df = df.applymap(lambda x: x.strip() if isinstance(x,str) else x)
        df.rename(columns={'Выберите свой муниципалитет (район)':'Муниципалитет'},inplace=True)

        # Исправляем файл с данными учащихся
        # ФИО
        part_fio_columns = ['Фамилия обучающегося', 'Имя обучающегося', 'Отчество обучающегося(при наличии)', 'ФИО законного представителя']  # колонки с типичными названиями
        df = prepare_fio_text_columns(df, part_fio_columns)  # очищаем колонки с фио

        # СНИЛС
        df['СНИЛС обучающегося'] = df['СНИЛС обучающегося'].apply(check_snils)
        dupl_snils = df[df['СНИЛС обучающегося'].duplicated(keep=False)]  # получаем дубликаты
        dupl_snils.to_excel(f'{result_folder}/Дубликаты по СНИЛС.xlsx',index=False)
        df = df.drop_duplicates(subset='СНИЛС обучающегося',keep='last') # оставляем только последнее вхождение
        # Паспортные данные
        prepared_columns_series_lst = ['Серия паспорта обучающегося','Серия паспорта законного представителя']
        prepared_columns_number_lst = ['Номер паспорта обучающегося','Номер паспорта законного представителя']
        df[prepared_columns_series_lst] = df[prepared_columns_series_lst].applymap(
            check_series_passport)  # обрабатываем серию паспорта

        df[prepared_columns_number_lst] = df[prepared_columns_number_lst].applymap(
            check_number_passport)  # обрабатываем номер паспорта

        df['Электронная почта законного представителя'] = df['Электронная почта законного представителя'].fillna('Ошибка: Не заполнено')
        df['Электронная почта законного представителя'] = df['Электронная почта законного представителя'].apply(
            lambda x: re.sub(r'\s', '', x) if x != 'Ошибка: Не заполнено' else x)

        # Обновляем индекс
        df.index = list(range(len(df)))

        # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками
        dct_sheet_error_df = dict()  # создаем словарь для хранения названия и датафрейма

        lst_name_columns = [name_cols for name_cols in df.columns if 'Unnamed' not in name_cols]  # получаем список колонок

        for idx, value in enumerate(lst_name_columns):
            # получаем ошибки
            temp_df = df[df[value].astype(str).str.contains('Ошибка')]  # фильтруем
            if temp_df.shape[0] == 0:
                continue

            temp_df = temp_df[value].to_frame()  # оставляем только одну колонку

            temp_df.insert(0, '№ строки с ошибкой в исходном файле', list(map(lambda x: x + 2, list(temp_df.index))))
            dct_sheet_error_df[value[:30]] = temp_df

        file_error_wb = write_df_to_excel_cheking_egisso(dct_sheet_error_df, write_index=False)
        file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        file_error_wb.save(f'{result_folder}/Ошибки {current_time}.xlsx')








        group_by = df.groupby(['Муниципалитет','Школа']).agg({'Фамилия обучающегося':'count'})
        group_by.rename(columns={'Фамилия обучающегося':'Количество зарегистрировавшихся'},inplace=True)

        dct_quota = {'Бичурский район':20,
                     'Заиграевский район':20,
                     'Закаменский район':30,
                     'Иволгинский район':30,
                     'Кабанский район':20,
                     'Кижингинский район':10,
                     'Курумканский район':30,
                     'Кяхтинский район':30,
                     'Мухоршибирский район':20,
                     'Селенгинский район':30,
                     'г. Северобайкальск':30,
                     'Хоринский район':10,
                     'Баунтовский (Эвенкийский) район':10,
                     'Прибайкальский район':10,
                     'Северобайкальский район':10,
                     'Тарбагатайский район':10,
                     'Баргузинский район':10,
                     'Джидинский район':10,
                     'Еравнинский район':10,
                     }


        quota_df = pd.DataFrame(columns=['Муниципалитет','Квота'])
        quota_df['Муниципалитет'] = dct_quota.keys()
        quota_df['Квота'] = dct_quota.values()

        group_mun = df.groupby(['Муниципалитет']).agg({'Фамилия обучающегося':'count'})
        group_mun.rename(columns={'Фамилия обучающегося':'Количество зарегистрировавшихся'},inplace=True)
        group_mun = group_mun.reset_index()

        quota_df = pd.merge(quota_df,group_mun,how='outer',left_on='Муниципалитет',right_on='Муниципалитет')
        quota_df.fillna(0,inplace=True)
        quota_df['Осталось до выполнения квоты'] = quota_df['Квота'] -quota_df['Количество зарегистрировавшихся']
        sum_row = quota_df.sum(axis=0,numeric_only=True)
        sum_row = sum_row.rename('Итого').to_frame().transpose()
        quota_df = pd.concat([quota_df,sum_row])
        quota_df.loc['Итого','Муниципалитет'] = 'Итого'
        with pd.ExcelWriter(f'{result_folder}/Сводка Первая профессия в {current_time}.xlsx') as writer:
            quota_df.to_excel(writer, sheet_name='Свод по квотам',index=False)
            group_by.to_excel(writer, sheet_name='Свод по школам')


        # Создание списков по муниципалитетам и списка для загрузки в мудл
        # создаем папку
        if not os.path.exists(f'{result_folder}/{"Списки по муниципалитетам"}'):
            os.makedirs(f'{result_folder}/{"Списки по муниципалитетам"}')

        moodle_df = df.copy() # копируем
        lst_username = [f'fp25_student{idx}' for idx in range(1,len(moodle_df)+1)]
        lst_password = [f'{random.randint(10000, 99999)}' for idx in range(1,len(moodle_df)+1)]
        lst_email = [f'fp25_student{idx}@mail.ru' for idx in range(1,len(moodle_df)+1)]
        moodle_df['Логин'] = lst_username
        moodle_df['Пароль'] = lst_password
        moodle_df['email'] = lst_email
        moodle_df['cohort1'] = moodle_df['Муниципалитет'].apply(create_cohort)
        moodle_df['ФИО'] = moodle_df['Фамилия обучающегося'] + ' ' + moodle_df['Имя обучающегося'] + ' '+ moodle_df['Отчество обучающегося(при наличии)']

        # Сохраняем файл для мудл
        out_moodle_df = moodle_df[['Логин','Пароль','Имя обучающегося','Фамилия обучающегося','email','cohort1']]
        out_moodle_df.columns = ['username','password','firstname','lastname','email','cohort1']
        out_moodle_df.to_excel(f'{result_folder}/Файл для MOODLE-{len(out_moodle_df)} строк.xlsx',index=False)

        # Сохраняем списки по муниципалитетам
        lst_value_column = moodle_df['Муниципалитет'].unique()

        for idx, value in enumerate(lst_value_column):
            wb = openpyxl.Workbook()  # создаем файл
            temp_df = moodle_df[moodle_df['Муниципалитет'] == value]  # отфильтровываем по значению
            temp_df = temp_df[['Школа','Класс','ФИО','Логин','Пароль','Номер телефона законного представителя']]
            # short_name = value[:40]  # получаем обрезанное значение
            # short_name = re.sub(r'[\r\b\n\t\'+()<> :"?*|\\/]', '_', short_name)
            for row in dataframe_to_rows(temp_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Устанавливаем автоширину для каждой колонки
            for column in wb['Sheet'].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb['Sheet'].column_dimensions[column_name].width = adjusted_width

            wb.save(f'{result_folder}/Списки по муниципалитетам/{value} учеников-{len(temp_df)}.xlsx')
            wb.close()


        # Сохраняем в формате для линди
        df = df.rename(columns={'Фамилия обучающегося':'Фамилия','Имя обучающегося':'Имя',
                           'Отчество обучающегося(при наличии)':'Отчество','Дата рождения обучающегося':'Дата_рождения',
                           'Пол обучающегося':'Пол','СНИЛС обучающегося':'СНИЛС',
                           'Гражданство обучающегося':'Гражданство','Фото СНИЛС обучающегося':'Фото_СНИЛС',
                           'Серия паспорта обучающегося':'Серия_паспорта','Номер паспорта обучающегося':'Номер_паспорта',
                           'Кем выдан паспорт обучающегося':'Кем_выдан_паспорт','Дата выдачи паспорта обучающегося':'Дата_выдачи_паспорта',
                           'ФИО законного представителя':'ФИО_представителя','Серия паспорта законного представителя':'Серия_паспорта_представителя',
                           'Номер паспорта законного представителя':'Номер_паспорта_представителя','Кем выдан паспорт законного представителя':'Кем_выдан_паспорт_представителя',
                           'Дата выдачи паспорта законного представителя':'Дата_выдачи_паспорта_представителя',
                           'Номер и серия свидетельства о рождении обучающегося':'Свидетельство_рождения',
                           'Дата выдачи свидетельства о рождении обучающегося':'Дата_выдачи_свидетельства',
                           'Номер телефона законного представителя':'Номер_телефона',
                           'Электронная почта законного представителя':'Электронная_почта',
                                })
        df['Номер_удостоверения'] = None
        df['Рег_номер'] = None
        df['Дата_выдачи'] = None
        df['Номер_договор'] = None
        df['Уровень_образования'] = None
        df['Фамилия_в_дипломе'] = None
        df['Серия_диплома'] = None
        df['Номер_диплома'] = None


        main_file_wb = write_df_error_egisso_to_excel({'Общий свод': df}, write_index=False)
        main_file_wb = del_sheet(main_file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        main_file_wb.save(f'{result_folder}/Список Первая профессия {len(df)}чел. в {current_time}.xlsx')

    except PermissionError as e:
        messagebox.showerror('Алора',
                             f'Закройте файлы созданные программой')
    except FileNotFoundError as e:
        messagebox.showerror('Алора',
                             f'Не удалось создать файл с названием {e}\n'
                             f'Уменьшите количество символов в соответствующей строке или выберите более короткий путь к итоговой папке')
    except PackageNotFoundError as e:
        messagebox.showerror('Алора',
                             f'Не удалось создать файл с названием {e}\n'
                             f'Уменьшите количество символов в соответствующей строке файла с данными в колонке по которой создаются имена файлов или выберите более короткий путь к итоговой папке')
    else:
        messagebox.showinfo('Алора', 'Создание документов успешно завершено !')


if __name__ == '__main__':
    main_data = 'data/Исходное Первая профессия.xlsx'
    main_result_folder = 'data/Результат'
    processing_data_first_prof(main_data,main_result_folder)


    print('Lindy Booth')


