"""
Скрипт для поиска ошибок в файлах выгрузок из СГО
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
import re
import time


def write_df_to_excel_error_prep_list(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            # для слишком длинных результатов
            if adjusted_width > 60:
                adjusted_width = 60
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

    return wb


def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb






def check_fio(value):
    if pd.isna(value):
        return 'Ошибка: Не заполнено'
    value = str(value)
    result = re.search(r'^[а-яёА-ЯЁ\-\— ]+$',value)
    if result:
        return value
    else:
        return f'Ошибка: допустимы только русские буквы дефис и тире: {value}'

def check_fio_patronomic(value):
    if pd.isna(value):
        return None
    value = str(value)
    result = re.search(r'^[а-яёА-ЯЁ\-\— ]+$',value)
    if result:
        return value
    else:
        return f'Ошибка: допустимы только русские буквы дефис и тире: {value}'


def check_date(value):
    if pd.isna(value):
        return 'Ошибка: Не заполнено'

    value = str(value)
    result = re.search(r'^\d\d\.\d\d\.\d{4}$',value)
    if result:
        return value
    else:
        return f'Ошибка: неправильная дата {value}'


def check_snils(value):
    if pd.isna(value):
        return 'Ошибка: Не заполнено'
    value = str(value)
    result = re.search(r'^\d{11}$',value)
    if result:
        return value
    else:
        return f'Ошибка: СНИЛС должен состоять из 11 цифр {value}'


def check_citizenship(value):
    if pd.isna(value):
        return 'Ошибка: Не заполнено'
    value = str(value)
    result = re.search(r'^[а-яёА-ЯЁ\-\— ]+$',value)
    if result:
        return value
    else:
        return f'Ошибка: допустимы только русские буквы дефис и тире: {value}'


def check_series_passport(row:pd.Series):
    """
    Функция для проверки серии паспорта, должно быть 4 цифры
    """
    type_doc, value = row.tolist()

    if type_doc == 'Иностранный паспорт':
        return value
    if type_doc == 'Вид на жительство':
        return value

    if pd.isna(value):
        return 'Ошибка: Не заполнено'
    value = str(value)
    result = re.search(r'^\d{4}$',value)
    if result:
        return value
    else:
        return f'Ошибка: Серия паспорта должна состоять из 4 цифр - {value}'

def check_number_passport(row:pd.Series):
    """
    Функция для проверки серии паспорта, должно быть 6 цифры
    """
    type_doc, value = row.tolist()

    if type_doc == 'Иностранный паспорт':
        if pd.isna(value):
            return 'Ошибка: Не заполнено'
        return value
    if type_doc == 'Вид на жительство':
        if pd.isna(value):
            return 'Ошибка: Не заполнено'
        return value
    if pd.isna(value):
        return 'Ошибка: Не заполнено'
    value = str(value)
    result = re.search(r'^\d{6}$',value)
    if result:
        return value
    else:
        return f'Ошибка: Номер паспорта должен состоять из 6 цифр - {value}'





def find_error_sgo(data_folder:str,end_folder:str):
    """
    Функция для поиска незаполненных ячеек в колонках Имя, Отчество, Дата рождения,СНИЛС, Гражданство, Серия документа, Номер документа
    :param data_folder:папка с данными
    :param end_folder:конечная папка с данными
    """
    # получаем время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    lst_cols = ['Фамилия','Имя','Отчество','Дата рождения','СНИЛС','Гражданство','Серия документа','Номер документа']
    for dirpath, dirnames, filenames in os.walk(data_folder):
        for file in filenames:
            if not file.startswith('~$') and (file.endswith('.csv')):
                name_file = file.split('.csv')[0].strip()
                print(name_file)
                df = pd.read_csv(f'{dirpath}/{file}',encoding='UTF-8',delimiter=';',dtype=str,on_bad_lines='warn')
                df['Фамилия'] = df['Фамилия'].apply(check_fio)
                df['Имя'] = df['Имя'].apply(check_fio)
                df['Отчество'] = df['Отчество'].apply(check_fio_patronomic)

                df['Дата рождения'] = df['Дата рождения'].apply(check_date)
                df['СНИЛС'] = df['СНИЛС'].apply(check_snils)
                df['Гражданство'] = df['Гражданство'].apply(check_citizenship)
                df['Серия документа'] = df[['Тип документа','Серия документа']].apply(check_series_passport,axis=1)
                df['Номер документа'] = df[['Тип документа','Номер документа']].apply(check_number_passport,axis=1)



                # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками
                dct_sheet_error_df = dict()  # создаем словарь для хранения названия и датафрейма
                used_name_sheet = set()  # множество для хранения значений которые уже были использованы


                for idx, value in enumerate(lst_cols):
                    # получаем ошибки
                    temp_df = df[df[value].astype(str).str.contains('Ошибка')]  # фильтруем
                    if temp_df.shape[0] == 0:
                        continue

                    # temp_df = temp_df[value].to_frame()  # оставляем только одну колонку

                    if 'БРПК' not in name_file:
                        temp_df.insert(0, '№ строки с ошибкой в исходном файле',
                                       list(map(lambda x: x + 2, list(temp_df.index))))
                    else:
                        temp_df.insert(0, '№ строки с ошибкой в исходном файле',
                                       list(map(lambda x: x + 3, list(temp_df.index))))
                    short_value = value[:27]  # получаем обрезанное значение
                    short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

                    if short_value.lower() in used_name_sheet:
                        short_value = f'{short_value}_{idx}'  # добавляем окончание
                    used_name_sheet.add(short_value.lower())

                    dct_sheet_error_df[short_value] = temp_df



                if len(dct_sheet_error_df) != 0:
                    inostr_df = df[
                        ~df['Гражданство'].str.contains('|'.join(['Россия', 'Ошибка']), case=False, regex=True)]
                    dct_sheet_error_df['Иностранцы'] = inostr_df
                    dct_sheet_error_df['Иностранцы по странам'] = inostr_df.groupby('Гражданство').agg(
                        {'Дата приказа': 'count'}).reset_index()
                    file_error_wb = write_df_to_excel_error_prep_list(dct_sheet_error_df, write_index=False)
                    file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                    file_error_wb.save(f'{end_folder}/Ошибки_{name_file}.xlsx')
                else:
                    inostr_df = df[
                        ~df['Гражданство'].str.contains('|'.join(['Россия', 'Ошибка']), case=False, regex=True)]
                    dct_sheet_error_df['Иностранцы'] = inostr_df
                    dct_sheet_error_df['Иностранцы по странам'] = inostr_df.groupby('Гражданство').agg(
                        {'Дата приказа': 'count'}).reset_index()
                    file_error_wb = write_df_to_excel_error_prep_list(dct_sheet_error_df, write_index=False)
                    file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                    file_error_wb.save(f'{end_folder}/НЕТ Ошибок_{name_file}.xlsx')













if __name__ == '__main__':
    main_data_folder = 'data/26_01_выгрузка студенты'
    main_end_folder = 'data/Ошибки'
    find_error_sgo(main_data_folder,main_end_folder)

    print('Lindy Booth !!!')




